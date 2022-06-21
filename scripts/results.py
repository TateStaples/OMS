import zipfile
import sys
import json
import openpyxl
import os


# parameter, percent change, # aircraft, launch time per run, For each sim: [AC->AC, HALO F-18, HALO E-2, P->AC, each color]
"""
Usage:
Create a folder with all of the simulations you want summarized
naming format should be:
- Base({number of aircraft}) [ie Base(25).sim] 
- {name}+-{percent change}({number of aircraft}) [ie HB-25(18).sim]
- {name}_{number of aircraft}_{'plus' or 'minus'}{percent change} [unchock_15_minus25.sim]

Once all your files are in one folder together run:
python3 results.py [path_to_folder_of_sims]

This will create a summary file in the same directory of the sims folder
"""

if len(sys.argv) > 1:
    # get file name from the first argument
    ROOT_PATH = sys.argv[1]
else:   # print usage and quit
    print("Usage: python3 summary.py [path_to_folder_of_sims]")
    quit()

save_path = "results.xlsx"
abbreviation = {"OrdnanceArmTime": 'Arm Ordanence Time', "HoldbackBarTime": 'Holdback Bar Time', "UnchockTime": 'Chocks and Chains', "SlipTripFallTime": "Slip Trip Fall Rate"}
base = dict(json.load(open("Setup.json")))


def get_workbook():
    if os.path.exists(save_path):
        wb = openpyxl.load_workbook(save_path)
        sheet = wb["Sheet"]
        start_index = sheet.max_row - 1
        print("adding to previous book")
    else:
        print("creating new book")
        wb = openpyxl.Workbook()
        sheet = wb["Sheet"]
        sheet.cell(row=1, column=4, value="Launch Times (minutes)")
        for index, col in enumerate(range(34, 34+11*30, 11), start=1):
            sheet.cell(row=1, column=col, value=f"Simulation {index}")
        header2 = ["Parameter", "Change percent", "Aircraft Count"] + [f"S{i}" for i in range(1, 31)] + \
                  ["Total Halo Violations (unexpected) A/C to A/C", "Total Halo Violations (Unexpected F-18s)", "Total Halo Violations (Unexpected E-2s)", "Halo Violations (unexpected) Person to A/C", "Brown Total Halo (unexpected) Violations", "Blue Total Halo (unexpected) Violations", "White Total Halo (unexpected) Violations", "Green Total Halo (unexpected) Violations", "Red Total Halo (unexpected) Violations", "Purple Total Halo (unexpected) Violations", "Yellow Total Halo (unexpected) Violations"] * 30 + \
                  [f"Mean Interval {i+1}" for i in range(30)]
        for col, val in enumerate(header2, start=1): sheet.cell(row=2, column=col, value=val)
        start_index = 1
    return sheet, wb, start_index


def read_res(run):
    launches = list()
    ac_to_person_violations = 0
    ac_to_ac_violations = 0
    F18_halo_violations = 0
    E2_halo_violations = 0
    Brown_Halo_Violations = 0
    Blue_Halo_Violations = 0
    White_Halo_Violations = 0
    Green_Halo_Violations = 0
    Red_Halo_Violations = 0
    Purple_Halo_Violations = 0
    Yellow_Halo_Violations = 0
    # stripped = (str(line).strip() for line in unzipped)  # make formatted rows
    lines = [line.split(b",") for line in run if line]  # split columns
    formatted_lines = list()
    for line in lines:
        new_line = list()
        for col in line:
            col = col.decode('utf-8')
            try:
                new_line.append(float(col))
            except:
                new_line.append(str(col).strip())
        formatted_lines.append(new_line)
    num_aircraft = int(formatted_lines[0][0].split('¡')[10])
    for row in formatted_lines:
        if row[0] == 2 and "Expected" not in row[3] and "ACtoAC" not in row[3]:
            ac_to_person_violations += 1
            if "D" in row[2]:
                Yellow_Halo_Violations += 1
            if "CnC" in row[2]:
                Blue_Halo_Violations += 1
            if "B" in row[2]:
                Brown_Halo_Violations += 1
            if "W" in row[2]:
                White_Halo_Violations += 1
            if "G" in row[2]:
                Green_Halo_Violations += 1
            if "R" in row[2]:
                Red_Halo_Violations += 1
            if "P" in row[2]:
                Purple_Halo_Violations += 1
        elif row[0] == 2 and "ACtoAC" in row[3] and "Expected" not in row[3]:
            ac_to_ac_violations += 1
            if row[1].startswith("F"):
                F18_halo_violations += 1
            elif row[1].startswith("E"):
                E2_halo_violations += 1
        elif row[0] == 3 and False:
            try:
                if row[1][0] == "F":
                    F18_time = row[2]
                if row[1][0] == "E":
                    E2_time = row[2]
            except:
                pass
        elif row[0] == 4:
            launches.append(row[1] * 0.001 / 60 - 0.75)

    halo_violation_packet = ac_to_ac_violations, F18_halo_violations, E2_halo_violations, ac_to_person_violations, Brown_Halo_Violations, Blue_Halo_Violations, White_Halo_Violations, Green_Halo_Violations, Red_Halo_Violations, Purple_Halo_Violations, Yellow_Halo_Violations
    return num_aircraft, halo_violation_packet, launches


def read_plb(run):
    end_times = list()
    prev_count = None
    for line in [l.decode('utf-8') for l in run]:
        sub_info = line.split(';')
        timestamp = float(sub_info[0])
        other_info = sub_info[1:]
        aircraft_info = list(filter(is_aircraft, other_info))
        count = len(aircraft_info)
        if prev_count is not None and count < prev_count:
            for i in range(prev_count-count):
                end_times.append(timestamp * 0.001 / 60)

        prev_count = count

        # for air_index, aircraft_string in enumerate(aircraft_info):
        #     type_, fuel, x, y, task, angle, wing_state, error = aircraft_string.split(",")  # see AircraftObject.cs
    # end_times.append(timestamp * 0.001 / 60 )
    return end_times


def read_json(data):
    for key in data:
        if data[key] != base[key]:
            actual_mean = base[key]["Mean"]
            mean = data[key]["Mean"]
            # print(key, actual_mean, mean)
            change = mean - actual_mean
            percent_change = change / actual_mean * 100
            parameter = abbreviation[key]
            break
    else:
        percent_change = 0
        parameter = "Base"
    return parameter, percent_change


def is_aircraft(info_string: str):  # see the simulationInstance.cs save method
    split = info_string.split(',')
    try:
        type_ = int(split[0])
        fuel = float(split[1])
        if type_ in (1, 2) and fuel > 1.0:
            return True
    except ValueError:
        return False
    return False


def scan_sims():
    all_results = list()
    for subdir, dirs, files in os.walk(ROOT_PATH):  # open target folder
        for sim_file in sorted(filter(lambda file: file.endswith(".sim"), files)):  # open each sim
            sim_name = sim_file.replace(".sim", "")
            print("loading", sim_name)
            sim_path = subdir + os.sep + sim_file
            with zipfile.ZipFile(sim_path, 'r') as zip_folder:  # unzip sim file
                launch_times = []
                halo_violations = []
                average_intervals = []

                for zip_file in zip_folder.infolist():  # open each run
                    if zip_file.filename.endswith("res"):
                        with zip_folder.open(zip_file) as run:  # unzip run
                            num_aircraft, h, launches = read_res(run)
                            halo_violations.append(h)
                            launch_times.extend(launches)
                    elif zip_file.filename.endswith("plb"):
                        with zip_folder.open(zip_file) as run:
                            end_times = read_plb(run)
                            intervals = list()
                            for i in range(len(end_times) - 1):
                                t1 = end_times[i]
                                t2 = end_times[i + 1]
                                delta = t2 - t1
                                intervals.append(delta)
                            mean_interval = sum(intervals)/len(intervals)
                            average_intervals.append(mean_interval)
                    elif zip_file.filename.endswith("json"):
                        data = json.load(zip_folder.open(zip_file))  # open(zip_file, 'r'))
                        parameter, percent_change = read_json(data)

                results = [parameter, percent_change, num_aircraft] + launch_times
                for packet in halo_violations:
                    results.extend(packet)
                results.extend(average_intervals)

                all_results.append(results)
                good_name = f"{parameter}_{int(percent_change)}({num_aircraft})"
                if good_name != sim_name:  # standardize the naming convention
                    os.rename(f"{ROOT_PATH}/{sim_name}.sim", f"{ROOT_PATH}/{good_name}.sim")
    return all_results


def write_results(results, sheet, start_index):
    parameters = sorted(list(set([result[0] for result in results])))
    counts = sorted(list(set([result[2] for result in results])))
    by_params = [[list() for j in range(len(counts))] for i in range(len(parameters))]
    index = start_index
    for result in results:
        by_params[parameters.index(result[0])][counts.index(result[2])].append(result)
    for type_ in by_params:
        for count in type_:
            count.sort(key=lambda x: x[1])
            for result in count:
                # print(result[1], result[2])
                for column, datum in enumerate(result, start=1):
                    sheet.cell(row=index+2, column=column, value=datum)
                index += 1


def save(wb):
    wb.save(save_path)
    print("saving to ", save_path)
    quit()


def main():
    sheet, wb, start_index = get_workbook()
    results = scan_sims()
    write_results(results, sheet, start_index)
    save(wb)


if __name__ == '__main__':
    main()
