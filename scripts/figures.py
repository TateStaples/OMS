import sys

import numpy as np
import openpyxl

"""
Usage:
Type python3 figures.py [path_to_summary_file] into your console (for example in the Google Drive the path would be oms_paper_data.xlsx)
After the program runs, it should create a new file called figures.xlsx that has the data organized into the sheets you need
"""
if len(sys.argv) > 1:
    # get file name from the first argument
    ROOT_PATH = sys.argv[1]
    # print("The arguments are: ", file_name)
else:   # print usage and quit
    ROOT_PATH = "Combined.xlsx"
    print("Usage: python3 figures.py [path_to_summary_file]")
    # quit()
# raw_data = "oms_paper_data.xlsx"
raw = openpyxl.load_workbook(ROOT_PATH)
raw = raw["Sheet"]

types = ["OAT", "unchock", "STF", "HB"]
percentages = ["-75.0", "-50.0", "-25.0", "-10.0", "10.0", "25.0", "50.0", "75.0"]
counts = [15, 20, 25]
colors = ("blue", "yellow", "brown", "green", "purple", "red", "white")

formatted_rows = list()
for row in raw.iter_rows():
    if row[0].value in (None, "Parameter"): continue
    row = [cell.value for cell in row]
    data = dict()
    data["type"] = row[0]
    data["variance"] = row[1]
    data["count"] = row[2]
    data["name"] = f"{data['type']}_{data['variance']}({data['count']})"
    data["times"] = np.array(row[3:33])
    data["launch time"] = np.mean(data["times"])
    data["launch std"] = np.std(data["times"])
    halo_data = list()
    for start_col in range(33, 32+11*30, 11):
        halo_violations = dict()
        halo_violations["AcAc"] = row[start_col]
        halo_violations["F-18"] = row[start_col + 1]
        halo_violations["E-2"] = row[start_col + 2]
        halo_violations["PAc"] = row[start_col + 3]
        halo_violations["brown"] = row[start_col + 4]
        halo_violations["blue"] = row[start_col + 5]
        halo_violations["white"] = row[start_col + 6]
        halo_violations["green"] = row[start_col + 7]
        halo_violations["red"] = row[start_col + 8]
        halo_violations["purple"] = row[start_col + 9]
        halo_violations["yellow"] = row[start_col + 10]
        halo_violations["total"] = sum(halo_violations[color] for color in colors) # sum(row[start_col+4:start_col+11])
        halo_data.append(halo_violations)
    data["halo data"] = halo_data
    data["halo averages"] = {key: sum(run[key]/30 for run in halo_data) for key in ("AcAc", "F-18", "E-2", "PAc", "brown", "blue", "white", "green", "red", "purple", "yellow", "total")}
    data["halo max"] = max(run["total"] for run in halo_data)
    formatted_rows.append(data)

wb = openpyxl.Workbook()
fig_2 = wb.create_sheet("Figure 2", 0)  # num aircraft vs departure rate (launch time / num aircraft)
fig_2.append(["Aircraft Configuration", "Departure Rate", "Observed Data", "Departure Std"])
for row in formatted_rows:
    if row["type"] != "Base": continue
    fig_2.append([row["count"], row["launch time"]/int(row["count"]), "", row["launch std"]/int(row["count"])])


fig_3 = wb.create_sheet("Figure 3")  # num aircraft vs unexpected halo violation percentages for each color
fig_3.append(["# Aircraft", "blue", "yellow", "brown", "green", "purple", "red", "white"])
max_halo_violations = max(row["halo averages"]["total"] for row in formatted_rows[:5])
for row in formatted_rows:
    if row["type"] != "Base": continue
    fig_3.append([row["count"]] + [row["halo averages"][color] / max_halo_violations for color in colors])

base = dict()
for row in formatted_rows:
    if row['type'] == "Base":
        base[row["count"]] = row["launch time"]
    else:
        if row["type"] not in wb.sheetnames:
            sheet = wb.create_sheet(row["type"])
            sheet.append([""] + counts)
            for i, percent in enumerate(percentages, start=2): sheet[f"A{i}"].value = percent
        else:
            sheet = wb.get_sheet_by_name(row["type"])
        variance = str(float(int(row["variance"])))
        row_index = percentages.index(variance)+2
        col_index = counts.index(row["count"]) + 2
        val = row["launch time"]/base[row["count"]]-1
        if row["type"] == "STF":
            print(row["launch time"], base[row["count"]])
        sheet.cell(row=row_index, column=col_index, value=val)
print(base)

fig_5 = wb.create_sheet("Figure 5")  # num aircraft vs launch time & total unexpected halo violations
# total halo violations too high
fig_5.append(["Aircraft Configuration", "Departure Rate", "Total Unexpected Halo Violations", "Departure Std"])
for row in formatted_rows:
    if row["type"] != "Base": continue
    fig_5.append([row["count"], row["launch time"], row["halo averages"]["total"], row["launch std"]])

fig_6 = wb.create_sheet("Figure 6")  # change % vs halo violation percent change
fig_6.append(["Percent Change in Parameter"] + types)
for i, percent in enumerate(percentages, start=2): fig_6[f"A{i}"].value = percent
for row in formatted_rows:
    if row['type'] == "Base":
        if int(row["count"]) == 25:
            base = row["halo averages"]["total"]
    elif row["type"] == "Holdback Bar Time": continue
    else:
        variance = str(float(int(row["variance"])))
        fig_6.cell(row=percentages.index(variance)+2, column=types.index(row["type"])+2, value=row["halo averages"]["total"]/base-1)


wb.save("figures.xlsx")