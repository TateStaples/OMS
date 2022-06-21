import zipfile
import sys
import statistics
import openpyxl
import os

"""
Usage:
Create a folder with all of the simulations you want summarized
naming format should be:
- Base({number of aircraft}) [ie Base(25).sim] 
- {name}+-{percent change}({number of aircraft}) [ie HB-25(18).sim]
- {name}_{number of aricraft}_{'plus' or 'minus'}{percent change} [unchock_15_minus25.sim]

Once all your files are in one folder together run:
python3 summary.py [path_to_folder_of_sims]

This will create a summary file in the same directory of the sims folder
"""

# ROOT_PATH = "/Users/22staples/Desktop/OAT"
if len(sys.argv)>1:
    # get file name from the first argument
    ROOT_PATH = sys.argv[1]
    # print("The arguments are: ", file_name)
else:   # print usage and quit
    print("Usage: python3 summary.py [path_to_folder_of_sims]")
    quit()
save_path = f"{ROOT_PATH.split('/')[-1]}_Summary.xlsx"
print(f"saving to {save_path}")


def Average(lst):
    return sum(lst) / len(lst) if len(lst) else None

wb = openpyxl.Workbook()
aggregate = wb.get_sheet_by_name("Sheet")
aggregate.title = "Aggregate"
aggregate.append(["Run Type",
                  "Total time of Launch",
                  "Average time per Aircraft",
                  "Average Time per F-18",
                  "Average time per E-2",
                  "Standard Deviation for F-18",
                  "Standard Deviation for E-2",
                  "Total Halo Violations (unexpected) A/C to A/C",
                  "Total Halo Violations (Unexpected F-18s)",
                  "Total Halo Violations (Unexpected E-2s)",
                  "Halo Violations (unexpected) Person to A/C",
                  "Brown Total Halo (unexpected) Violations",
                  "Blue Total Halo (unexpected) Violations",
                  "White Total Halo (unexpected) Violations",
                  "Green Total Halo (unexpected) Violations",
                  "Red Total Halo (unexpected) Violations",
                  "Purple Total Halo (unexpected) Violations",
                  "Yellow Total Halo (unexpected) Violations"])

for subdir, dirs, files in os.walk(ROOT_PATH):
    for sim_index, sim_file in enumerate(sorted(filter(lambda file: file.endswith(".sim"), files))):
        sim_name = sim_file.replace(".sim", "")
        sim_path = subdir + os.sep + sim_file
        with zipfile.ZipFile(sim_path, 'r') as zip_folder:
            Total_Times = []
            F18_Averages = []
            E2_Averages = []
            All_Averages = []
            F18_StdDevs = []
            E2_StdDevs = []
            AC_to_Person_unexpected_violations = []
            AC_to_AC_unexpected_violations = []
            F18_Halo_Violations = []
            E2_Halo_Violations = []
            All_Brown_Halo_Violations = []
            All_Blue_Halo_Violations = []
            All_White_Halo_Violations = []
            All_Green_Halo_Violations = []
            All_Red_Halo_Violations = []
            All_Purple_Halo_Violations = []
            All_Yellow_Halo_Violations = []
            Individual_F18_Times = []
            Individual_E2_Times = []

            for zip_file in filter(lambda file: file.filename.endswith("res"), zip_folder.infolist()):
                with zip_folder.open(zip_file) as run:
                    F18_Times = []
                    E2_Times = []
                    ac_to_person_violations = 0
                    ac_to_ac_violations = 0
                    F18_halo_violations = 0
                    E2_halo_violations = 0
                    Brown_Halo_Violations = 0
                    Blue_Halo_Violations = 0
                    White_Halo_Violations = 0
                    Green_Halo_Violations = 0
                    Red_Halo_Violations = 0
                    Purple_Halo_Violations = 0
                    Yellow_Halo_Violations = 0
                    # stripped = (str(line).strip() for line in unzipped)  # make formatted rows
                    lines = [line.split(b",") for line in run if line]  # split columns
                    formatted_lines = list()
                    for line in lines:
                        new_line = list()
                        for col in line:
                            col = col.decode('utf-8')
                            try:
                                new_line.append(float(col))
                            except:
                                new_line.append(str(col).strip())
                        formatted_lines.append(new_line)
                    for row in formatted_lines:
                        # print(row)
                        if row[0] == 2 and "Expected" not in row[3] and "ACtoAC" not in row[3]:
                            ac_to_person_violations += 1
                            if "D" in row[2]:
                                Yellow_Halo_Violations += 1
                            if "CnC" in row[2]:
                                Blue_Halo_Violations += 1
                            if "B" in row[2]:
                                Brown_Halo_Violations += 1
                            if "W" in row[2]:
                                White_Halo_Violations += 1
                            if "G" in row[2]:
                                Green_Halo_Violations += 1
                            if "R" in row[2]:
                                Red_Halo_Violations += 1
                            if "P" in row[2]:
                                Purple_Halo_Violations += 1
                        elif row[0] == 2 and "ACtoAC" in row[3]:
                            ac_to_ac_violations += 1
                            if row[1].startswith("F"):
                                F18_halo_violations += 1
                            elif row[1].startswith("E"):
                                E2_halo_violations += 1
                        elif row[0] == 3:
                            try:
                                if row[1][0] == "F":
                                    #                                 print ("Appending {} to F18".format(row[2]))
                                    F18_Times.append(row[2])
                                    Individual_F18_Times.append(row[2])
                                if row[1][0] == "E":
                                    #                                 print ("Appending {} to E2".format(row[2]))
                                    E2_Times.append(row[2])
                                    Individual_E2_Times.append(row[2])
                            except:
                                print("skipping ", row[0])
                        elif row[0] == 4:
                            Total_Times.append(row[1] * 0.001 / 60 - 0.75)

                    F18_Averages.append(Average(F18_Times))
                    E2_Averages.append(Average(E2_Times))
                    All_Averages.append((Average(F18_Times) + Average(E2_Times))/2)
                    F18_StdDevs.append(statistics.stdev(F18_Times))
                    try:
                        E2_StdDevs.append(statistics.stdev(E2_Times))
                    except:
                        E2_StdDevs.append(E2_Times[0])
                    AC_to_Person_unexpected_violations.append(ac_to_person_violations)
                    AC_to_AC_unexpected_violations.append(ac_to_ac_violations)
                    F18_Halo_Violations.append(F18_halo_violations)
                    E2_Halo_Violations.append(E2_halo_violations)
                    All_Yellow_Halo_Violations.append(Yellow_Halo_Violations)
                    All_Blue_Halo_Violations.append(Blue_Halo_Violations)
                    All_Brown_Halo_Violations.append(Brown_Halo_Violations)
                    All_White_Halo_Violations.append(White_Halo_Violations)
                    All_Green_Halo_Violations.append(Green_Halo_Violations)
                    All_Red_Halo_Violations.append(Red_Halo_Violations)
                    All_Purple_Halo_Violations.append(Purple_Halo_Violations)

            index = sim_index
            info_cell = aggregate.cell(row=index+2, column=1, value=sim_name)
            total_time_cell = aggregate.cell(row=index + 2, column=2, value=Average(Total_Times))
            average_current_cell = aggregate.cell(row=index + 2, column=3, value=Average(All_Averages))
            F18_current_cell = aggregate.cell(row=index + 2, column=4, value=Average(F18_Times))
            E2_current_cell = aggregate.cell(row=index + 2, column=5, value=Average(E2_Times))
            F18_stdDev_cell = aggregate.cell(row=index + 2, column=6, value=Average(F18_StdDevs))
            E2_stdDev_cell = aggregate.cell(row=index + 2, column=7, value=Average(E2_StdDevs))
            total_ac_to_ac_violations_cell = aggregate.cell(row=index + 2, column=8, value=Average(AC_to_AC_unexpected_violations))
            F18_unexpected_violations = aggregate.cell(row=index + 2, column=9, value=Average(F18_Halo_Violations))
            E2_unexpected_violations = aggregate.cell(row=index + 2, column=10, value=Average(E2_Halo_Violations))
            total_ac_to_person_violations_cell = aggregate.cell(row=index + 2, column=11, value=Average(AC_to_Person_unexpected_violations))
            All_Brown_Halo_Violations_cell = aggregate.cell(row=index + 2, column=12, value=Average(All_Brown_Halo_Violations))
            All_Blue_Halo_Violations_cell = aggregate.cell(row=index + 2, column=13, value=Average(All_Blue_Halo_Violations))
            All_White_Halo_Violations_cell = aggregate.cell(row=index + 2, column=14, value=Average(All_White_Halo_Violations))
            All_Green_Halo_Violations_cell = aggregate.cell(row=index + 2, column=15, value=Average(All_Green_Halo_Violations))
            All_Red_Halo_Violations_cell = aggregate.cell(row=index + 2, column=16, value=Average(All_Red_Halo_Violations))
            All_Purple_Halo_Violations_cell = aggregate.cell(row=index + 2, column=17, value=Average(All_Purple_Halo_Violations))
            All_Yellow_Halo_Violations_cell = aggregate.cell(row=index + 2, column=18, value=Average(All_Yellow_Halo_Violations))
wb.save(save_path)
quit()


if __name__ == '__main__':
    pass